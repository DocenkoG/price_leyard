# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                      # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, quoted, dump_cell, currencyType, openX, sheetByName
import csv
import requests, lxml.html



def nameToId(value) :
    result = ''
    for ch in value:
        if (ch != " " and ch != "/" and ch != "\\" and ch != '_' and ch != "," and 
            ch != "'" and ch != "." and ch != "-" and ch != "!" and ch != "@" and 
            ch != "#" and ch != "$" and ch != "%" and ch != "^" and ch != "&" and 
            ch != "*" and ch != "(" and ch != ")" and ch != "[" and ch != "]" and 
            ch != "{" and ch != ":" and ch != '"' and ch != ";"                  ) :
            result = result + ch

    length = len(result)
    if  length > 50 :
        point = int(length/2)
        result = result[:13] + result[point-12:point+13] + result[-12: ]
    
    return result



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]-1
        if item in ('закупка','продажа','цена1') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh).find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена2','цена1') :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('Call for Pricing') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_excel2csv(cfg):
    csvFName  = cfg.get('basic','filename_out')
    priceFName= cfg.get('basic','filename_in')
    #sheetName = cfg.get('basic','sheetname')
    
    log.debug('Reading file ' + priceFName )
    book = openX(priceFName)
    sheetNames = book.sheetnames
    print(sheetNames)

    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)

    outFileEUR = open( csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriterEUR = csv.DictWriter(outFileEUR, fieldnames=out_cols )
    csvWriterEUR.writeheader()

    pricelines   =[]
    for sheetName in sheetNames:
        log.debug("Sheet   "+sheetName)
        sheet = book[sheetName]
    
        recOut  ={}
        grp = ''
        subgrp1 = ''
        subgrp2 = ''
        for i in range(2, sheet.max_row +1) :                                # xlsx
    #   for i in range(2, sheet.nrows) :                                     # xls
            i_last = i
            try:
                impValues = getXlsxString(sheet, i, in_cols_j)               # xlsx
                #impValues = getXlsString(sheet, i, in_cols_j)               # xls
                #print( impValues )
                if impValues['группа_'] != '':                                                 # Группа
                    grp = impValues['группа_']
                    subgrp1 = ''
                    subgrp2 = ''
                    continue
                if impValues['подгруппа1'] != '':                                             # Подгруппа 1
                    subgrp1 = impValues['подгруппа1']
                    subgrp2 = ''
                    continue
                if impValues['подгруппа2'] != '':                                             # Подгруппа 1
                    subgrp2 = "/" + impValues['подгруппа2']
                    continue
                if impValues['цена1']=='0': # (ccc.value == None) or (ccc2.value == None) :   # Пустая строка
                    #print( 'Пустая строка. i=',i, impValues )
                    continue
                else :                                                                        # Обычная строка
                    impValues['группа_'] = grp
                    impValues['подгруппа2'] = subgrp2
                    impValues['подгруппа1'] = subgrp1
                    for outColName in out_template.keys() :
                        shablon = out_template[outColName]
                        for key in impValues.keys():
                            if shablon.find(key) >= 0 :
                                shablon = shablon.replace(key, impValues[key])
                        if (outColName in('закупка','продажа')) and ('*' in shablon) :
                            p = shablon.find("*")
                            vvv1 = float(shablon[:p])
                            vvv2 = float(shablon[p+1:])
                            shablon = str(round(vvv1 * vvv2, 2))
                        elif (outColName=='код') :
                            shablon = nameToId(shablon)
                        recOut[outColName] = shablon.strip()
    
                    pricelines.append(dict(recOut))
                    #csvWriterEUR.writerow(recOut)    
                
            except Exception as e:
                print(e)
                if str(e) == "'NoneType' object has no attribute 'rgb'":
                    pass
                else:
                    log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )
        log.info('Обработано ' +str(i_last)+ ' строк.')

    pricelines.sort(key=lambda recOut: recOut['код'])
    j = 0
    for k in range(1, len(pricelines)):
        if pricelines[k]['код'] in(pricelines[k-1]['код']):
            j += 1
            log.info('--- Дубликат ' + str(j) )
            log.info( pricelines[k-1]['код'] + '.' + pricelines[k-1]['код производителя'] + ' ' + pricelines[k-1]['описание'])
            log.info( pricelines[k]['код'] + '.' +  pricelines[k]['код производителя']+ ' ' + pricelines[k]['описание'])
            pricelines[k]['код'] = pricelines[k]['код'] + '.' + pricelines[k]['код производителя']
    for line_ in pricelines:
        csvWriterEUR.writerow(line_)
        
    outFileEUR.close()



def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    if price_datetime+qty_seconds < time.time() :
        file_age = round((time.time()-price_datetime)/24/60/60)
        log.error('Файл "'+fileName+'" устарел!  Допустимый период '+ str(qty_days)+' дней, а ему ' + str(file_age) )
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def processing(cfgFName):
    log.info('----------------------- Processing '+cfgFName )
    cfg = config_read(cfgFName)
    filename_out = cfg.get('basic','filename_out')
    filename_in  = cfg.get('basic','filename_in')
    
    convert_excel2csv(cfg)
    if os.name == 'nt' :
        folderName = os.path.basename(os.getcwd())
        foutRUR = filename_out[:-4]+'_RUR'+filename_out[-4:]
        foutEUR = filename_out[:-4]+'_EUR'+filename_out[-4:]
        foutUSD = filename_out[:-4]+'_USD'+filename_out[-4:]


def main( dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый 
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          '+dealerName )

    if  os.path.exists('getting.cfg'):     
        cfg = config_read('getting.cfg')
        filename_new = cfg.get('basic','filename_new')
        
        rc_download = False
        if cfg.has_section('download'):
            rc_download = download(cfg)
        if rc_download==True or is_file_fresh( filename_new, int(cfg.get('basic','срок годности'))):
            pass
        else:
            return
        
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            processing(cfgFName)


if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
